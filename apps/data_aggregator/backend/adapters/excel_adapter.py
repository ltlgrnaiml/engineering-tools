"""Excel file adapter for DAT.

Per ADR-0012: Profile-Driven Extraction & AdapterFactory Pattern.
Per SPEC-0026: Adapter Interface & Registry specification.

This adapter handles Excel file formats (.xlsx, .xls) using Polars
for efficient data processing.

Supported formats:
- XLSX (Excel 2007+) - .xlsx
- XLS (Legacy Excel) - .xls

Features:
- Multi-sheet support with sheet selection
- Column type inference
- Header row detection

Note: Excel adapter does NOT support streaming due to file format limitations.
For large Excel files, consider converting to CSV first.
"""

import asyncio
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, AsyncIterator

import polars as pl

from shared.contracts.dat.adapter import (
    AdapterCapabilities,
    AdapterError,
    AdapterErrorCode,
    AdapterMetadata,
    BaseFileAdapter,
    ColumnInfo,
    FileValidationResult,
    InferredDataType,
    ReadOptions,
    ReadResult,
    SchemaProbeResult,
    SheetInfo,
    StreamChunk,
    StreamOptions,
    ValidationIssue,
    ValidationSeverity,
)

__version__ = "1.0.0"


def _polars_dtype_to_inferred(dtype: pl.DataType) -> InferredDataType:
    """Convert Polars dtype to InferredDataType enum.

    Args:
        dtype: Polars data type.

    Returns:
        Corresponding InferredDataType enum value.
    """
    dtype_str = str(dtype).lower()

    if "int" in dtype_str:
        return InferredDataType.INTEGER
    elif "float" in dtype_str or "decimal" in dtype_str:
        return InferredDataType.FLOAT
    elif "bool" in dtype_str:
        return InferredDataType.BOOLEAN
    elif "datetime" in dtype_str:
        return InferredDataType.DATETIME
    elif "date" in dtype_str:
        return InferredDataType.DATE
    elif "time" in dtype_str:
        return InferredDataType.TIME
    elif "binary" in dtype_str:
        return InferredDataType.BINARY
    elif "null" in dtype_str:
        return InferredDataType.NULL
    elif "str" in dtype_str or "utf8" in dtype_str:
        return InferredDataType.STRING
    else:
        return InferredDataType.UNKNOWN


class ExcelAdapter(BaseFileAdapter):
    """Adapter for Excel files (.xlsx, .xls).

    Uses Polars for efficient data processing. Supports multi-sheet files
    with explicit sheet selection.

    Note: Streaming is NOT supported for Excel files due to format limitations.
    For large Excel files (> 100MB), consider converting to CSV first.

    Attributes:
        _metadata: Cached adapter metadata.

    Example:
        >>> adapter = ExcelAdapter()
        >>> result = await adapter.probe_schema("data.xlsx")
        >>> df, read_result = await adapter.read_dataframe(
        ...     "data.xlsx",
        ...     options=ReadOptions(extra={"sheet_name": "Sheet2"})
        ... )
    """

    def __init__(self) -> None:
        """Initialize the Excel adapter."""
        self._metadata = AdapterMetadata(
            adapter_id="excel",
            name="Excel Adapter",
            version=__version__,
            file_extensions=[".xlsx", ".xls"],
            mime_types=[
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "application/vnd.ms-excel",
            ],
            capabilities=AdapterCapabilities(
                supports_streaming=False,  # Excel doesn't support streaming
                supports_schema_inference=True,
                supports_random_access=False,
                supports_column_selection=True,
                max_recommended_file_size_mb=100,  # Large Excel files are slow
                supported_compressions=[],  # Excel has its own compression
                supports_multiple_sheets=True,
            ),
            description="Parse Excel files (.xlsx, .xls) with multi-sheet support",
            author="system",
            icon="file-spreadsheet",
        )

    @property
    def metadata(self) -> AdapterMetadata:
        """Return adapter metadata for registry.

        Returns:
            AdapterMetadata with adapter ID, capabilities, etc.
        """
        return self._metadata

    async def probe_schema(
        self,
        file_path: str,
        options: ReadOptions | None = None,
    ) -> SchemaProbeResult:
        """Probe Excel file to discover schema without reading all data.

        Args:
            file_path: Relative path to the Excel file.
            options: Optional read options (sheet_name in extras).

        Returns:
            SchemaProbeResult with column info and sheet list.

        Raises:
            AdapterError: If file cannot be probed.
        """
        start_time = datetime.now(timezone.utc)
        options = options or ReadOptions()
        errors: list[str] = []
        warnings: list[str] = []

        try:
            path = Path(file_path)
            if not path.exists():
                raise AdapterError(
                    code=AdapterErrorCode.FILE_NOT_FOUND,
                    message=f"File not found: {file_path}",
                    file_path=file_path,
                    adapter_id=self._metadata.adapter_id,
                )

            file_size = path.stat().st_size

            # Get sheet names
            def _get_sheet_names() -> list[str]:
                # Polars doesn't have a direct way to list sheets
                # We'll use openpyxl or xlrd depending on format
                if path.suffix.lower() == ".xlsx":
                    try:
                        from openpyxl import load_workbook
                        wb = load_workbook(path, read_only=True, data_only=True)
                        names = wb.sheetnames
                        wb.close()
                        return names
                    except ImportError:
                        # Fall back to reading with polars (will get first sheet only)
                        return ["Sheet1"]
                else:
                    # For .xls files
                    try:
                        import xlrd
                        wb = xlrd.open_workbook(str(path))
                        return wb.sheet_names()
                    except ImportError:
                        return ["Sheet1"]

            sheet_names = await asyncio.to_thread(_get_sheet_names)

            # Determine which sheet to probe
            sheet_name = options.extra.get("sheet_name")
            sheet_index = options.extra.get("sheet_index", 0)

            if sheet_name and sheet_name not in sheet_names:
                warnings.append(f"Sheet '{sheet_name}' not found, using first sheet")
                sheet_name = sheet_names[0] if sheet_names else None

            target_sheet = sheet_name or (sheet_names[sheet_index] if sheet_names else None)

            # Read sample for schema inference
            sample_rows = min(options.infer_schema_length, 1000)

            def _read_sample() -> pl.DataFrame:
                read_kwargs: dict[str, Any] = {
                    "source": path,
                    "sheet_name": target_sheet,
                    "read_options": {"n_rows": sample_rows},
                    "infer_schema_length": sample_rows,
                }
                return pl.read_excel(**read_kwargs)

            df = await asyncio.to_thread(_read_sample)

            # Build column info
            columns: list[ColumnInfo] = []
            for i, col_name in enumerate(df.columns):
                col = df[col_name]
                dtype = col.dtype

                sample_values = col.head(10).to_list()
                null_count = col.null_count()

                try:
                    distinct_count = col.n_unique()
                except Exception:
                    distinct_count = None

                columns.append(
                    ColumnInfo(
                        name=col_name,
                        position=i,
                        inferred_type=_polars_dtype_to_inferred(dtype),
                        nullable=null_count > 0,
                        sample_values=sample_values,
                        null_count=null_count,
                        distinct_count_estimate=distinct_count,
                    )
                )

            # Build sheet info
            sheets: list[SheetInfo] = []
            for idx, name in enumerate(sheet_names):
                sheets.append(
                    SheetInfo(
                        sheet_name=name,
                        sheet_index=idx,
                        row_count_estimate=len(df) if name == target_sheet else None,
                        column_count=len(df.columns) if name == target_sheet else None,
                        is_empty=False,
                    )
                )

            end_time = datetime.now(timezone.utc)
            duration_ms = (end_time - start_time).total_seconds() * 1000

            return SchemaProbeResult(
                file_path=file_path,
                file_size_bytes=file_size,
                adapter_id=self._metadata.adapter_id,
                columns=columns,
                row_count_estimate=len(df),
                row_count_exact=False,
                encoding_detected=None,  # Excel handles encoding internally
                delimiter_detected=None,
                has_header_row=True,
                sheets=sheets,
                compression_detected=None,
                probed_at=start_time,
                probe_duration_ms=duration_ms,
                sample_rows_read=len(df),
                errors=errors,
                warnings=warnings,
            )

        except AdapterError:
            raise
        except Exception as e:
            raise AdapterError(
                code=AdapterErrorCode.SCHEMA_INFERENCE_FAILED,
                message=f"Failed to probe Excel schema: {e}",
                file_path=file_path,
                adapter_id=self._metadata.adapter_id,
                details={"error": str(e)},
            ) from e

    async def read_dataframe(
        self,
        file_path: str,
        options: ReadOptions | None = None,
    ) -> tuple[pl.DataFrame, ReadResult]:
        """Read Excel file into a Polars DataFrame.

        Args:
            file_path: Relative path to the Excel file.
            options: Read options. Supports:
                - extra.sheet_name: Sheet name to read
                - extra.sheet_index: Sheet index to read (default: 0)
                - columns: Columns to include
                - row_limit: Maximum rows to read

        Returns:
            Tuple of (DataFrame, ReadResult metadata).

        Raises:
            AdapterError: If file cannot be read.
        """
        start_time = datetime.now(timezone.utc)
        options = options or ReadOptions()
        warnings: list[str] = []

        try:
            path = Path(file_path)
            if not path.exists():
                raise AdapterError(
                    code=AdapterErrorCode.FILE_NOT_FOUND,
                    message=f"File not found: {file_path}",
                    file_path=file_path,
                    adapter_id=self._metadata.adapter_id,
                )

            file_size = path.stat().st_size

            # Get sheet selection
            sheet_name = options.extra.get("sheet_name")
            sheet_index = options.extra.get("sheet_index", 0)

            def _read_excel() -> pl.DataFrame:
                read_kwargs: dict[str, Any] = {
                    "source": path,
                }

                if sheet_name:
                    read_kwargs["sheet_name"] = sheet_name
                elif sheet_index is not None:
                    read_kwargs["sheet_id"] = sheet_index + 1  # 1-indexed in polars

                # Read options
                read_opts: dict[str, Any] = {}
                if options.row_limit:
                    read_opts["n_rows"] = options.row_limit + options.skip_rows

                if read_opts:
                    read_kwargs["read_options"] = read_opts

                df = pl.read_excel(**read_kwargs)

                # Apply skip_rows
                if options.skip_rows > 0:
                    df = df.slice(options.skip_rows)

                # Apply row_limit
                if options.row_limit and len(df) > options.row_limit:
                    df = df.head(options.row_limit)

                # Select columns if specified
                if options.columns:
                    available = set(df.columns)
                    selected = [c for c in options.columns if c in available]
                    if selected:
                        df = df.select(selected)

                # Exclude columns if specified
                if options.exclude_columns:
                    cols_to_keep = [
                        c for c in df.columns if c not in options.exclude_columns
                    ]
                    df = df.select(cols_to_keep)

                return df

            df = await asyncio.to_thread(_read_excel)

            end_time = datetime.now(timezone.utc)
            duration_ms = (end_time - start_time).total_seconds() * 1000

            was_truncated = options.row_limit is not None and len(df) >= options.row_limit

            result = ReadResult(
                file_path=file_path,
                adapter_id=self._metadata.adapter_id,
                rows_read=len(df),
                columns_read=len(df.columns),
                bytes_read=file_size,
                read_duration_ms=duration_ms,
                warnings=warnings,
                was_truncated=was_truncated,
            )

            return df, result

        except AdapterError:
            raise
        except Exception as e:
            raise AdapterError(
                code=AdapterErrorCode.PARSE_ERROR,
                message=f"Failed to read Excel file: {e}",
                file_path=file_path,
                adapter_id=self._metadata.adapter_id,
                details={"error": str(e)},
            ) from e

    async def stream_dataframe(
        self,
        file_path: str,
        options: StreamOptions | None = None,
    ) -> AsyncIterator[tuple[pl.DataFrame, StreamChunk]]:
        """Stream is NOT supported for Excel files.

        Excel files cannot be efficiently streamed due to their binary format.
        For large Excel files, consider converting to CSV first.

        Args:
            file_path: Path to the file.
            options: Stream options (ignored).

        Raises:
            AdapterError: Always raises with STREAMING_NOT_SUPPORTED code.
        """
        raise AdapterError(
            code=AdapterErrorCode.STREAMING_NOT_SUPPORTED,
            message=(
                "Excel files do not support streaming due to binary format limitations. "
                "For large Excel files (> 100MB), consider converting to CSV first."
            ),
            file_path=file_path,
            adapter_id=self._metadata.adapter_id,
            recoverable=False,
        )
        # Yield statement to make this a generator (never reached)
        yield  # type: ignore[misc]

    async def validate_file(
        self,
        file_path: str,
        options: ReadOptions | None = None,
    ) -> FileValidationResult:
        """Validate Excel file can be processed by this adapter.

        Args:
            file_path: Relative path to the Excel file.
            options: Optional read options for validation context.

        Returns:
            FileValidationResult with validation status and issues.
        """
        start_time = datetime.now(timezone.utc)
        issues: list[ValidationIssue] = []

        try:
            path = Path(file_path)

            # Check file exists
            if not path.exists():
                issues.append(
                    ValidationIssue(
                        severity=ValidationSeverity.ERROR,
                        code="FILE_NOT_FOUND",
                        message=f"File does not exist: {file_path}",
                        suggestion="Check the file path and ensure the file exists.",
                    )
                )
                return self._build_validation_result(file_path, start_time, issues)

            # Check file is not empty
            file_size = path.stat().st_size
            if file_size == 0:
                issues.append(
                    ValidationIssue(
                        severity=ValidationSeverity.ERROR,
                        code="EMPTY_FILE",
                        message="File is empty",
                        suggestion="Provide a non-empty Excel file.",
                    )
                )
                return self._build_validation_result(file_path, start_time, issues)

            # Check file size warning
            if file_size > 100 * 1024 * 1024:  # > 100MB
                issues.append(
                    ValidationIssue(
                        severity=ValidationSeverity.WARNING,
                        code="LARGE_FILE",
                        message=f"File is large ({file_size / 1024 / 1024:.1f} MB). "
                        "Processing may be slow.",
                        suggestion="Consider converting to CSV for better performance.",
                    )
                )

            # Try to read the file header
            def _validate_content() -> list[ValidationIssue]:
                content_issues: list[ValidationIssue] = []
                try:
                    # Try to read just 1 row to validate format
                    pl.read_excel(path, read_options={"n_rows": 1})
                except Exception as e:
                    error_msg = str(e).lower()
                    if "password" in error_msg or "encrypted" in error_msg:
                        content_issues.append(
                            ValidationIssue(
                                severity=ValidationSeverity.ERROR,
                                code="PASSWORD_PROTECTED",
                                message="File appears to be password protected",
                                suggestion="Remove password protection before processing.",
                            )
                        )
                    elif "corrupt" in error_msg or "invalid" in error_msg:
                        content_issues.append(
                            ValidationIssue(
                                severity=ValidationSeverity.ERROR,
                                code="CORRUPT_FILE",
                                message=f"File appears to be corrupt: {e}",
                                suggestion="Try opening the file in Excel and re-saving it.",
                            )
                        )
                    else:
                        content_issues.append(
                            ValidationIssue(
                                severity=ValidationSeverity.ERROR,
                                code="INVALID_FORMAT",
                                message=f"Cannot read Excel file: {e}",
                            )
                        )
                return content_issues

            content_issues = await asyncio.to_thread(_validate_content)
            issues.extend(content_issues)

            return self._build_validation_result(file_path, start_time, issues)

        except Exception as e:
            issues.append(
                ValidationIssue(
                    severity=ValidationSeverity.ERROR,
                    code="VALIDATION_FAILED",
                    message=f"Validation failed: {e}",
                )
            )
            return self._build_validation_result(file_path, start_time, issues)

    def _build_validation_result(
        self,
        file_path: str,
        start_time: datetime,
        issues: list[ValidationIssue],
    ) -> FileValidationResult:
        """Build FileValidationResult from issues list.

        Args:
            file_path: Path to the validated file.
            start_time: When validation started.
            issues: List of validation issues found.

        Returns:
            FileValidationResult with computed fields.
        """
        end_time = datetime.now(timezone.utc)
        duration_ms = (end_time - start_time).total_seconds() * 1000

        error_count = sum(
            1 for i in issues if i.severity == ValidationSeverity.ERROR
        )
        warning_count = sum(
            1 for i in issues if i.severity == ValidationSeverity.WARNING
        )

        return FileValidationResult(
            file_path=file_path,
            adapter_id=self._metadata.adapter_id,
            is_valid=error_count == 0,
            issues=issues,
            error_count=error_count,
            warning_count=warning_count,
            validated_at=start_time,
            validation_duration_ms=duration_ms,
        )
